from fastapi import APIRouter, UploadFile, File
from fastapi.responses import JSONResponse, FileResponse
import pandas as pd
import shutil, os, json, io, csv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import Counter
import chardet
import uuid
from pathlib import Path
from typing import Optional
import threading
from datetime import datetime, timedelta

router = APIRouter(tags=["Report Difference Analyzer"])

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

progress_lock = threading.Lock()
ANALYSIS_PROGRESS = {"phase": "idle", "percent": 0, "message": "Waiting..."}
FILE_RETENTION_HOURS = 24


# ----------------------------------------------------------
# 🧩 Utility Functions
# ----------------------------------------------------------
def log(msg: str):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


def update_progress(phase: str, percent: int, message: str):
    global ANALYSIS_PROGRESS
    with progress_lock:
        ANALYSIS_PROGRESS = {"phase": phase, "percent": percent, "message": message}
        log(f"Progress: {phase} {percent}% - {message}")


def cleanup_old_files():
    """Auto-delete older uploaded files."""
    try:
        cutoff = datetime.now() - timedelta(hours=FILE_RETENTION_HOURS)
        for item in Path(UPLOAD_DIR).iterdir():
            if item.is_file() and datetime.fromtimestamp(item.stat().st_mtime) < cutoff:
                item.unlink(missing_ok=True)
    except Exception:
        pass


def save_uploaded_file(upload: UploadFile, target_path: str):
    """Save uploaded file and return its path."""
    upload.file.seek(0)
    with open(target_path, "wb") as out_f:
        shutil.copyfileobj(upload.file, out_f)
    upload.file.seek(0)
    return target_path


def get_file_extension(upload: UploadFile) -> str:
    """Extract file extension from uploaded file."""
    if upload.filename:
        ext = Path(upload.filename).suffix.lower()
        # Default to .xlsx if no extension found
        return ext if ext else '.xlsx'
    return '.xlsx'


def detect_encoding(file_path: str) -> str:
    """Detect file encoding using chardet."""
    try:
        with open(file_path, "rb") as f:
            raw = f.read(50000)
        detected = chardet.detect(raw)
        return detected.get("encoding") or "utf-8"
    except Exception:
        return "utf-8"


import unicodedata
import csv
from pathlib import Path
from typing import Optional

def read_saved_table(path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """
    Robust reader for Excel and CSV:
      - Detects Excel vs CSV by extension AND magic bytes
      - For Excel: loads header=None first, finds first non-empty header row, reloads with header
      - For CSV: tries multiple encodings, sniffs delimiter, loads with engine='python'
      - Cleans header names (removes invisible chars, BOMs, normalizes Unicode)
      - DOES NOT drop duplicate ROWS (that caused your row-count shrinkage)
      - Logs debug info about columns so you can see exactly what was read
    """
    ext = Path(path).suffix.lower()
    try:
        # Read file header bytes
        with open(path, "rb") as f:
            start_bytes = f.read(8)

        # ----------------------------------------------------------
        # ✅ Detect Excel files by magic number ("PK" header for XLSX)
        # ----------------------------------------------------------
        is_excel = (start_bytes[:2] == b"PK" or b"xl/" in start_bytes or 
                   ext in ['.xlsx', '.xlsm', '.xltx', '.xltm', '.xls'])
        
        if is_excel:
            from openpyxl import load_workbook
            wb = load_workbook(filename=path, read_only=True, data_only=True)
            sheet = wb[sheet_name] if sheet_name else wb.active

            rows = list(sheet.iter_rows(values_only=True))
            wb.close()

            # Detect header row
            header_row_index = None
            for i, row in enumerate(rows[:10]):
                if row and any(cell not in (None, "", " ") for cell in row):
                    header_row_index = i
                    break

            if header_row_index is None:
                raise ValueError("No valid header row found in Excel sheet")

            headers = [str(c).strip() if c else "" for c in rows[header_row_index]]
            headers = [
                unicodedata.normalize("NFKC", h)
                .replace("\ufeff", "")
                .replace("\u200b", "")
                .strip()
                for h in headers
            ]

            valid_indices = [i for i, h in enumerate(headers) if h.strip()]
            headers = [headers[i] for i in valid_indices]

            data_rows = [
                [row[i] if i < len(row) else None for i in valid_indices]
                for row in rows[header_row_index + 1 :]
            ]

            df = pd.DataFrame(data_rows, columns=headers)

        # ----------------------------------------------------------
        # ✅ Otherwise treat as CSV
        # ----------------------------------------------------------
        else:
            encodings = ["utf-8", "utf-8-sig", "latin1", "cp1252"]
            df = None
            last_error = None
            for enc in encodings:
                try:
                    with open(path, "r", encoding=enc, errors="replace") as f:
                        sample = f.read(4096)
                    try:
                        delimiter = csv.Sniffer().sniff(sample).delimiter
                    except Exception:
                        delimiter = ","
                    df = pd.read_csv(
                        path,
                        dtype=str,
                        keep_default_na=False,
                        encoding=enc,
                        delimiter=delimiter,
                        engine="python",
                    )
                    break
                except Exception as e:
                    last_error = e
                    continue

            if df is None:
                raise ValueError(f"Failed to read CSV file with any encoding. Last error: {last_error}")

        # ----------------------------------------------------------
        # 🧹 Normalize headers
        # ----------------------------------------------------------
        df.columns = (
            df.columns.astype(str)
            .str.strip()
            .str.replace("\ufeff", "", regex=False)
            .str.replace("\u200b", "", regex=False)
            .str.replace(r"[\n\r\t]+", " ", regex=True)
        )
        df.columns = [unicodedata.normalize("NFKC", c) for c in df.columns]
        df = df.loc[:, ~df.columns.duplicated()]

        print(f"[DEBUG] ✅ Loaded '{os.path.basename(path)}' → Rows: {len(df)}, Cols: {len(df.columns)}")
        for i, col in enumerate(df.columns[:15]):
            print(f"   {i+1}. {repr(col)}")
        if len(df.columns) > 15:
            print(f"   ... +{len(df.columns) - 15} more columns")

        return df

    except Exception as e:
        raise ValueError(f"Failed to read table from '{path}': {e}")

# ----------------------------------------------------------
# 🧾 List Excel Sheets
# ----------------------------------------------------------
@router.post("/sheets/")
async def list_excel_sheets(file: UploadFile = File(...)):
    """Return available sheet names for uploaded Excel file."""
    try:
        file.file.seek(0)
        wb = load_workbook(filename=file.file, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        file.file.seek(0)
        return {"sheets": sheets}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=400)


# ----------------------------------------------------------
# 🧮 Analyze Reports (Row-level comparison)
# ----------------------------------------------------------
@router.post("/analyze/")
async def analyze_reports(
    reference: UploadFile = File(...),
    main: UploadFile = File(...),
    ref_sheet: Optional[str] = None,
    main_sheet: Optional[str] = None,
    strict_mode: bool = False
):
    """Full comparison with missing, extra, and mismatched rows."""
    session = str(uuid.uuid4())[:8]
    try:
        cleanup_old_files()
        update_progress("starting", 5, "Saving files...")

        # Get proper file extensions from uploaded files
        ref_ext = get_file_extension(reference)
        main_ext = get_file_extension(main)
        
        # Save with proper extensions
        ref_path = os.path.join(UPLOAD_DIR, f"ref_{session}{ref_ext}")
        main_path = os.path.join(UPLOAD_DIR, f"main_{session}{main_ext}")
        save_uploaded_file(reference, ref_path)
        save_uploaded_file(main, main_path)

        update_progress("reading", 15, "Reading data files...")
        ref_df = read_saved_table(ref_path, sheet_name=ref_sheet)
        main_df = read_saved_table(main_path, sheet_name=main_sheet)

        def normalize(cols):
            return (
                cols.astype(str)
                .str.strip()
                .str.replace("\ufeff", "", regex=False)
                .str.replace("\u200b", "", regex=False)
                .str.replace(r"\s+", " ", regex=True)
                .str.lower()
            )

        ref_df.columns = normalize(ref_df.columns)
        main_df.columns = normalize(main_df.columns)

        # Remove duplicate column headers
        ref_df = ref_df.loc[:, ~ref_df.columns.duplicated()]
        main_df = main_df.loc[:, ~main_df.columns.duplicated()]

        ref_cols = list(ref_df.columns)
        main_cols = list(main_df.columns)

        # Column validation
        missing_cols = [c for c in ref_cols if c not in main_cols]
        extra_cols = [c for c in main_cols if c not in ref_cols]
        if missing_cols or extra_cols:
            update_progress("error", 100, "Column mismatch detected.")
            return JSONResponse({
                "error": "Column mismatch",
                "missing_columns": missing_cols,
                "extra_columns": extra_cols
            }, status_code=400)

        shared_cols = [c for c in ref_cols if c in main_cols]
        if not shared_cols:
            return JSONResponse({"error": "No shared columns"}, status_code=400)

        def norm_cell(x):
            if pd.isna(x) or x is None:
                return ""
            s = str(x).strip()
            
            if not strict_mode:
                # Normalize whitespace
                s = " ".join(s.split())
                
                # Remove trailing .0 from numbers (e.g., "4219.0" -> "4219")
                if s.endswith('.0') and s[:-2].replace('.', '').replace('-', '').isdigit():
                    s = s[:-2]
                
                # Try to normalize as number if possible
                try:
                    # Check if it's a number
                    float_val = float(s)
                    # If it's a whole number, remove decimals
                    if float_val == int(float_val):
                        s = str(int(float_val))
                    else:
                        # Keep as float but normalize format
                        s = str(float_val)
                except (ValueError, OverflowError):
                    # Not a number, keep as string
                    pass
            
            return s

        ref_df_proc = ref_df[shared_cols].fillna("").astype(str).applymap(norm_cell)
        main_df_proc = main_df[shared_cols].fillna("").astype(str).applymap(norm_cell)

        update_progress("comparing", 45, "Comparing rows...")
        # Remove completely empty rows (all cells blank)
        ref_df_proc = ref_df_proc.dropna(how="all")
        main_df_proc = main_df_proc.dropna(how="all")

        # Remove duplicates for clean comparison
        ref_df_proc = ref_df_proc.drop_duplicates()
        main_df_proc = main_df_proc.drop_duplicates()
        
        # 🔄 SMART MATCHING: Instead of sorting, find best row matches
        update_progress("matching", 50, "Finding matching rows...")
        log(f"📊 Ref rows: {len(ref_df_proc)}, Main rows: {len(main_df_proc)}")
        
        ref_tuples = [tuple(r) for r in ref_df_proc.values.tolist()]
        main_tuples = [tuple(r) for r in main_df_proc.values.tolist()]
        
        ref_counter = Counter(ref_tuples)
        main_counter = Counter(main_tuples)
        
        # Find exact matches first
        exact_matches = {}
        main_used = set()
        
        for ref_idx, ref_tuple in enumerate(ref_tuples):
            for main_idx, main_tuple in enumerate(main_tuples):
                if main_idx not in main_used and ref_tuple == main_tuple:
                    exact_matches[ref_idx] = main_idx
                    main_used.add(main_idx)
                    break
        
        # For remaining rows, find "close" matches (rows that differ in only a few cells)
        ref_unmatched = [i for i in range(len(ref_tuples)) if i not in exact_matches]
        main_unmatched = [i for i in range(len(main_tuples)) if i not in main_used]
        
        close_matches = {}
        similarity_threshold = len(shared_cols) * 0.7  # At least 70% columns must match
        
        for ref_idx in ref_unmatched:
            best_match = None
            best_similarity = 0
            
            for main_idx in main_unmatched:
                # Count matching cells
                matching_cells = sum(1 for k in range(len(shared_cols)) 
                                   if ref_tuples[ref_idx][k] == main_tuples[main_idx][k])
                
                if matching_cells > best_similarity and matching_cells >= similarity_threshold:
                    best_similarity = matching_cells
                    best_match = main_idx
            
            if best_match is not None:
                close_matches[ref_idx] = best_match
                main_used.add(best_match)
                main_unmatched.remove(best_match)
        
        # Combine exact and close matches
        all_matches = {**exact_matches, **close_matches}
        matched_count = len(all_matches)
        
        log(f"✅ Found {len(exact_matches)} exact matches, {len(close_matches)} close matches")

        ref_tuples = [tuple(r) for r in ref_df_proc.values.tolist()]
        main_tuples = [tuple(r) for r in main_df_proc.values.tolist()]

        ref_counter = Counter(ref_tuples)
        main_counter = Counter(main_tuples)

        missing_rows, extra_rows = [], []

        # Rows in reference but not matched
        missing_row_indices = []
        for i, ref_tuple in enumerate(ref_tuples):
            if i not in all_matches:
                missing_rows.append(ref_tuple)
                missing_row_indices.append(i)
        
        # Rows in main but not matched
        extra_row_indices = []
        for i, main_tuple in enumerate(main_tuples):
            if i not in main_used:
                extra_rows.append(main_tuple)
                extra_row_indices.append(i)
        
        log(f"📊 DETAILED ANALYSIS:")
        log(f"   Total Ref Rows: {len(ref_tuples)}")
        log(f"   Total Main Rows: {len(main_tuples)}")
        log(f"   Matched: {matched_count}")
        log(f"   Missing (in Ref, not in Main): {len(missing_rows)} rows")
        log(f"   Missing indices: {missing_row_indices[:10]}..." if len(missing_row_indices) > 10 else f"   Missing indices: {missing_row_indices}")
        log(f"   Extra (in Main, not in Ref): {len(extra_rows)} rows")
        log(f"   Extra indices: {extra_row_indices[:10]}..." if len(extra_row_indices) > 10 else f"   Extra indices: {extra_row_indices}")

        missing_path = os.path.join(UPLOAD_DIR, f"missing_{session}.xlsx")
        extra_path = os.path.join(UPLOAD_DIR, f"extra_{session}.xlsx")
        mismatch_path = os.path.join(UPLOAD_DIR, f"mismatch_{session}.xlsx")

        # Delete old session files if they exist (prevent serving stale data)
        for old_file in [missing_path, extra_path, mismatch_path]:
            if os.path.exists(old_file):
                try:
                    os.remove(old_file)
                    log(f"🗑️ Deleted old file: {os.path.basename(old_file)}")
                except Exception as e:
                    log(f"⚠️ Could not delete old file {old_file}: {e}")

        # Only create files if there's data to save
        missing_file_created = False
        extra_file_created = False
        mismatch_file_created = False

        # Create Missing Rows file ONLY if there are missing rows
        if len(missing_rows) > 0:
            log(f"🔍 MISSING ROWS CHECK:")
            log(f"   missing_rows list length: {len(missing_rows)}")
            log(f"   missing_row_indices length: {len(missing_row_indices)}")
            
            if len(missing_row_indices) > 0:
                df_missing = ref_df_proc.iloc[missing_row_indices].copy()
                log(f"   df_missing shape: {df_missing.shape}")
                df_missing.insert(0, 'Original Row Number', [i + 2 for i in missing_row_indices])
                df_missing.to_excel(missing_path, index=False)
                missing_file_created = True
                log(f"✅ Created missing rows file: {missing_path}")
                log(f"   File size: {os.path.getsize(missing_path)} bytes")
            else:
                log(f"⚠️ Missing rows count mismatch: list has {len(missing_rows)} but indices has {len(missing_row_indices)}")
        else:
            log(f"ℹ️ No missing rows - skipping file creation")
            
        # Create Extra Rows file ONLY if there are extra rows
        if len(extra_rows) > 0:
            log(f"🔍 EXTRA ROWS CHECK:")
            log(f"   extra_rows list length: {len(extra_rows)}")
            log(f"   extra_row_indices length: {len(extra_row_indices)}")
            
            if len(extra_row_indices) > 0:
                df_extra = main_df_proc.iloc[extra_row_indices].copy()
                log(f"   df_extra shape: {df_extra.shape}")
                df_extra.insert(0, 'Original Row Number', [i + 2 for i in extra_row_indices])
                df_extra.to_excel(extra_path, index=False)
                extra_file_created = True
                log(f"✅ Created extra rows file: {extra_path}")
                log(f"   File size: {os.path.getsize(extra_path)} bytes")
            else:
                log(f"⚠️ Extra rows count mismatch: list has {len(extra_rows)} but indices has {len(extra_row_indices)}")
        else:
            log(f"ℹ️ No extra rows - skipping file creation")

        update_progress("analyzing", 70, "Analyzing cell-level mismatches...")
        mismatches = []
        mismatch_rows = []
        
        # Check matched rows for cell-level differences
        for ref_idx, main_idx in all_matches.items():
            row_has_mismatch = False
            row_mismatches = []
            for col_idx, col in enumerate(shared_cols):
                ref_val = ref_df_proc.iat[ref_idx, col_idx]
                main_val = main_df_proc.iat[main_idx, col_idx]
                if ref_val != main_val:
                    row_has_mismatch = True
                    row_mismatches.append({
                        "ref_row": ref_idx,
                        "main_row": main_idx,
                        "column": col,
                        "reference_value": ref_val,
                        "main_value": main_val
                    })
            
            if row_has_mismatch:
                mismatch_rows.append((ref_idx, main_idx))
                mismatches.extend(row_mismatches)
        
        log(f"🔍 Found {len(mismatches)} cell mismatches in {len(mismatch_rows)} row pairs")
        
        # Create Cell Mismatches file ONLY if there are actual mismatches
        if len(mismatches) > 0 and len(mismatch_rows) > 0:
                # Create complete row comparison with ALL columns
                combined_data = []
                
                log(f"📝 Creating mismatch report for {len(mismatch_rows)} row pairs")
                
                for ref_idx, main_idx in mismatch_rows:
                    # Add a row identifier with actual Excel row numbers
                    ref_row = {
                        "Source": "REFERENCE", 
                        "Ref Excel Row": ref_idx + 2,  # +2 because Excel is 1-based and has header
                        "Main Excel Row": main_idx + 2
                    }
                    main_row = {
                        "Source": "MAIN", 
                        "Ref Excel Row": ref_idx + 2, 
                        "Main Excel Row": main_idx + 2
                    }
                    
                    # Add all columns from both reference and main
                    for col in shared_cols:
                        ref_val = ref_df_proc.iat[ref_idx, ref_df_proc.columns.get_loc(col)]
                        main_val = main_df_proc.iat[main_idx, main_df_proc.columns.get_loc(col)]
                        ref_row[col] = ref_val
                        main_row[col] = main_val
                    
                    combined_data.append(ref_row)
                    combined_data.append(main_row)
                    
                    # Add empty row for visual separation
                    combined_data.append({col: "" for col in ["Source", "Ref Excel Row", "Main Excel Row"] + shared_cols})
                
                df_mismatch = pd.DataFrame(combined_data)
                df_mismatch.to_excel(mismatch_path, index=False)
                log(f"📁 Created mismatch file with {len(mismatch_rows)} row pairs")

                try:
                    wb = load_workbook(mismatch_path)
                    ws = wb.active
                    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                    # Build a map of which cells differ for each row pair
                    mismatch_map = {}
                    for m in mismatches:
                        key = (m["ref_row"], m["main_row"])
                        col = m["column"]
                        if key not in mismatch_map:
                            mismatch_map[key] = set()
                        mismatch_map[key].add(col)

                    # Color the cells: red for mismatches, green for matches
                    row_num = 2  # Start after header
                    for ref_idx, main_idx in mismatch_rows:
                        mismatched_cols = mismatch_map.get((ref_idx, main_idx), set())
                        
                        # Color REFERENCE row
                        for col_idx, col_name in enumerate(["Source", "Ref Excel Row", "Main Excel Row"] + shared_cols, start=1):
                            cell = ws.cell(row=row_num, column=col_idx)
                            if col_name in mismatched_cols:
                                cell.fill = red
                            elif col_name not in ["Source", "Ref Excel Row", "Main Excel Row"]:
                                cell.fill = green
                            else:
                                cell.fill = blue
                        
                        row_num += 1
                        
                        # Color MAIN row
                        for col_idx, col_name in enumerate(["Source", "Ref Excel Row", "Main Excel Row"] + shared_cols, start=1):
                            cell = ws.cell(row=row_num, column=col_idx)
                            if col_name in mismatched_cols:
                                cell.fill = red
                            elif col_name not in ["Source", "Ref Excel Row", "Main Excel Row"]:
                                cell.fill = green
                            else:
                                cell.fill = blue
                        
                        row_num += 2  # Skip separator row

                    wb.save(mismatch_path)
                    wb.close()
                    log(f"✅ Applied color formatting to mismatch file")
                except Exception as e:
                    log(f"⚠️ Failed to color mismatches: {e}")
                
                mismatch_file_created = True
        else:
            log(f"ℹ️ No cell mismatches - skipping file creation")

        # Clean up temporary files
        try:
            if os.path.exists(ref_path):
                os.remove(ref_path)
            if os.path.exists(main_path):
                os.remove(main_path)
            log(f"🧹 Cleaned up temporary upload files")
        except Exception as e:
            log(f"⚠️ Failed to cleanup temp files: {e}")

        update_progress("done", 100, "Analysis complete ✅")

        log(f"📦 FINAL FILE STATUS:")
        log(f"   missing_file_created: {missing_file_created}")
        log(f"   extra_file_created: {extra_file_created}")
        log(f"   mismatch_file_created: {mismatch_file_created}")
        log(f"   Missing path to return: {os.path.basename(missing_path) if missing_file_created else 'None'}")
        log(f"   Extra path to return: {os.path.basename(extra_path) if extra_file_created else 'None'}")
        log(f"   Mismatch path to return: {os.path.basename(mismatch_path) if mismatch_file_created else 'None'}")

        return {
            "status": "success",
            "session": session,
            "rowsCompared": len(ref_df_proc),
            "sharedColumns": len(shared_cols),
            "matchedRows": matched_count,
            "missingRows": len(missing_rows),
            "extraRows": len(extra_rows),
            "cellMismatches": len(mismatches),
            "missingPath": os.path.basename(missing_path) if missing_file_created else None,
            "extraPath": os.path.basename(extra_path) if extra_file_created else None,
            "mismatchPath": os.path.basename(mismatch_path) if mismatch_file_created else None,
        }

    except Exception as e:
        update_progress("error", 100, f"Error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


# ----------------------------------------------------------
# 🧩 Debug Samples
# ----------------------------------------------------------
@router.post("/debug/")
async def debug_samples(reference: UploadFile = File(...), main: UploadFile = File(...), sample_size: int = 10):
    """Return a small normalized sample for debugging."""
    try:
        # Get proper file extensions
        ref_ext = get_file_extension(reference)
        main_ext = get_file_extension(main)
        
        ref_path = os.path.join(UPLOAD_DIR, f"debug_ref_{uuid.uuid4().hex[:6]}{ref_ext}")
        main_path = os.path.join(UPLOAD_DIR, f"debug_main_{uuid.uuid4().hex[:6]}{main_ext}")
        save_uploaded_file(reference, ref_path)
        save_uploaded_file(main, main_path)

        ref_df = read_saved_table(ref_path)
        main_df = read_saved_table(main_path)

        ref_df = ref_df.fillna("").astype(str)
        main_df = main_df.fillna("").astype(str)
        shared_cols = [c for c in ref_df.columns if c in main_df.columns]

        return {
            "reference_sample": ref_df[shared_cols].head(sample_size).values.tolist(),
            "main_sample": main_df[shared_cols].head(sample_size).values.tolist(),
            "shared_columns": shared_cols
        }
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


# ----------------------------------------------------------
# 🟢 Progress Tracker
# ----------------------------------------------------------
@router.get("/progress")
async def get_progress():
    with progress_lock:
        return ANALYSIS_PROGRESS.copy()


# ----------------------------------------------------------
# 📦 Download Endpoints
# ----------------------------------------------------------
def get_latest_file(prefix: str) -> Optional[str]:
    files = [f for f in os.listdir(UPLOAD_DIR) if f.startswith(prefix)]
    if not files:
        return None
    files.sort(key=lambda x: os.path.getmtime(os.path.join(UPLOAD_DIR, x)), reverse=True)
    return os.path.join(UPLOAD_DIR, files[0])


@router.get("/download-missing/")
def download_missing(session: str = "latest"):
    path = get_latest_file("missing_") if session == "latest" else os.path.join(UPLOAD_DIR, f"missing_{session}.xlsx")
    if not path or not os.path.exists(path):
        return JSONResponse({"error": "Missing rows file not found"}, status_code=404)
    # Verify file is not empty
    if os.path.getsize(path) == 0:
        return JSONResponse({"error": "Missing rows file is empty"}, status_code=404)
    return FileResponse(path, filename="missing_rows.xlsx")


@router.get("/download-extra/")
def download_extra(session: str = "latest"):
    path = get_latest_file("extra_") if session == "latest" else os.path.join(UPLOAD_DIR, f"extra_{session}.xlsx")
    if not path or not os.path.exists(path):
        return JSONResponse({"error": "Extra rows file not found"}, status_code=404)
    # Verify file is not empty
    if os.path.getsize(path) == 0:
        return JSONResponse({"error": "Extra rows file is empty"}, status_code=404)
    return FileResponse(path, filename="extra_rows.xlsx")


@router.get("/download-mismatched/")
def download_mismatched(session: str = "latest"):
    path = get_latest_file("mismatch_") if session == "latest" else os.path.join(UPLOAD_DIR, f"mismatch_{session}.xlsx")
    if not path or not os.path.exists(path):
        return JSONResponse({"error": "Mismatched rows file not found"}, status_code=404)
    # Verify file is not empty
    if os.path.getsize(path) == 0:
        return JSONResponse({"error": "Mismatched rows file is empty"}, status_code=404)
    return FileResponse(path, filename="mismatched_rows.xlsx")
# from fastapi import APIRouter, UploadFile, File
# from fastapi.responses import JSONResponse, FileResponse
# import pandas as pd
# import shutil, os, json, io, csv
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill
# from collections import Counter
# import chardet
# import uuid
# from pathlib import Path
# from typing import Optional
# import threading
# from datetime import datetime, timedelta

# router = APIRouter(tags=["Report Difference Analyzer"])

# UPLOAD_DIR = "uploads"
# os.makedirs(UPLOAD_DIR, exist_ok=True)

# progress_lock = threading.Lock()
# ANALYSIS_PROGRESS = {"phase": "idle", "percent": 0, "message": "Waiting..."}
# FILE_RETENTION_HOURS = 24


# # ----------------------------------------------------------
# # 🧩 Utility Functions
# # ----------------------------------------------------------
# def log(msg: str):
#     print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


# def update_progress(phase: str, percent: int, message: str):
#     global ANALYSIS_PROGRESS
#     with progress_lock:
#         ANALYSIS_PROGRESS = {"phase": phase, "percent": percent, "message": message}
#         log(f"Progress: {phase} {percent}% - {message}")


# def cleanup_old_files():
#     """Auto-delete older uploaded files."""
#     try:
#         cutoff = datetime.now() - timedelta(hours=FILE_RETENTION_HOURS)
#         for item in Path(UPLOAD_DIR).iterdir():
#             if item.is_file() and datetime.fromtimestamp(item.stat().st_mtime) < cutoff:
#                 item.unlink(missing_ok=True)
#     except Exception:
#         pass


# def save_uploaded_file(upload: UploadFile, target_path: str):
#     """Save uploaded file and return its path."""
#     upload.file.seek(0)
#     with open(target_path, "wb") as out_f:
#         shutil.copyfileobj(upload.file, out_f)
#     upload.file.seek(0)
#     return target_path


# def get_file_extension(upload: UploadFile) -> str:
#     """Extract file extension from uploaded file."""
#     if upload.filename:
#         ext = Path(upload.filename).suffix.lower()
#         # Default to .xlsx if no extension found
#         return ext if ext else '.xlsx'
#     return '.xlsx'


# def detect_encoding(file_path: str) -> str:
#     """Detect file encoding using chardet."""
#     try:
#         with open(file_path, "rb") as f:
#             raw = f.read(50000)
#         detected = chardet.detect(raw)
#         return detected.get("encoding") or "utf-8"
#     except Exception:
#         return "utf-8"


# import unicodedata
# import csv
# from pathlib import Path
# from typing import Optional

# def read_saved_table(path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
#     """
#     Robust reader for Excel and CSV:
#       - Detects Excel vs CSV by extension AND magic bytes
#       - For Excel: loads header=None first, finds first non-empty header row, reloads with header
#       - For CSV: tries multiple encodings, sniffs delimiter, loads with engine='python'
#       - Cleans header names (removes invisible chars, BOMs, normalizes Unicode)
#       - DOES NOT drop duplicate ROWS (that caused your row-count shrinkage)
#       - Logs debug info about columns so you can see exactly what was read
#     """
#     ext = Path(path).suffix.lower()
#     try:
#         # Read file header bytes
#         with open(path, "rb") as f:
#             start_bytes = f.read(8)

#         # ----------------------------------------------------------
#         # ✅ Detect Excel files by magic number ("PK" header for XLSX)
#         # ----------------------------------------------------------
#         is_excel = (start_bytes[:2] == b"PK" or b"xl/" in start_bytes or 
#                    ext in ['.xlsx', '.xlsm', '.xltx', '.xltm', '.xls'])
        
#         if is_excel:
#             from openpyxl import load_workbook
#             wb = load_workbook(filename=path, read_only=True, data_only=True)
#             sheet = wb[sheet_name] if sheet_name else wb.active

#             rows = list(sheet.iter_rows(values_only=True))
#             wb.close()

#             # Detect header row
#             header_row_index = None
#             for i, row in enumerate(rows[:10]):
#                 if row and any(cell not in (None, "", " ") for cell in row):
#                     header_row_index = i
#                     break

#             if header_row_index is None:
#                 raise ValueError("No valid header row found in Excel sheet")

#             headers = [str(c).strip() if c else "" for c in rows[header_row_index]]
#             headers = [
#                 unicodedata.normalize("NFKC", h)
#                 .replace("\ufeff", "")
#                 .replace("\u200b", "")
#                 .strip()
#                 for h in headers
#             ]

#             valid_indices = [i for i, h in enumerate(headers) if h.strip()]
#             headers = [headers[i] for i in valid_indices]

#             data_rows = [
#                 [row[i] if i < len(row) else None for i in valid_indices]
#                 for row in rows[header_row_index + 1 :]
#             ]

#             df = pd.DataFrame(data_rows, columns=headers)

#         # ----------------------------------------------------------
#         # ✅ Otherwise treat as CSV
#         # ----------------------------------------------------------
#         else:
#             encodings = ["utf-8", "utf-8-sig", "latin1", "cp1252"]
#             df = None
#             last_error = None
#             for enc in encodings:
#                 try:
#                     with open(path, "r", encoding=enc, errors="replace") as f:
#                         sample = f.read(4096)
#                     try:
#                         delimiter = csv.Sniffer().sniff(sample).delimiter
#                     except Exception:
#                         delimiter = ","
#                     df = pd.read_csv(
#                         path,
#                         dtype=str,
#                         keep_default_na=False,
#                         encoding=enc,
#                         delimiter=delimiter,
#                         engine="python",
#                     )
#                     break
#                 except Exception as e:
#                     last_error = e
#                     continue

#             if df is None:
#                 raise ValueError(f"Failed to read CSV file with any encoding. Last error: {last_error}")

#         # ----------------------------------------------------------
#         # 🧹 Normalize headers
#         # ----------------------------------------------------------
#         df.columns = (
#             df.columns.astype(str)
#             .str.strip()
#             .str.replace("\ufeff", "", regex=False)
#             .str.replace("\u200b", "", regex=False)
#             .str.replace(r"[\n\r\t]+", " ", regex=True)
#         )
#         df.columns = [unicodedata.normalize("NFKC", c) for c in df.columns]
#         df = df.loc[:, ~df.columns.duplicated()]

#         print(f"[DEBUG] ✅ Loaded '{os.path.basename(path)}' → Rows: {len(df)}, Cols: {len(df.columns)}")
#         for i, col in enumerate(df.columns[:15]):
#             print(f"   {i+1}. {repr(col)}")
#         if len(df.columns) > 15:
#             print(f"   ... +{len(df.columns) - 15} more columns")

#         return df

#     except Exception as e:
#         raise ValueError(f"Failed to read table from '{path}': {e}")

# # ----------------------------------------------------------
# # 🧾 List Excel Sheets
# # ----------------------------------------------------------
# @router.post("/sheets/")
# async def list_excel_sheets(file: UploadFile = File(...)):
#     """Return available sheet names for uploaded Excel file."""
#     try:
#         file.file.seek(0)
#         wb = load_workbook(filename=file.file, read_only=True, data_only=True)
#         sheets = wb.sheetnames
#         wb.close()
#         file.file.seek(0)
#         return {"sheets": sheets}
#     except Exception as e:
#         return JSONResponse({"error": str(e)}, status_code=400)


# # ----------------------------------------------------------
# # 🧮 Analyze Reports (Row-level comparison)
# # ----------------------------------------------------------
# @router.post("/analyze/")
# async def analyze_reports(
#     reference: UploadFile = File(...),
#     main: UploadFile = File(...),
#     ref_sheet: Optional[str] = None,
#     main_sheet: Optional[str] = None,
#     strict_mode: bool = False
# ):
#     """Full comparison with missing, extra, and mismatched rows."""
#     session = str(uuid.uuid4())[:8]
#     try:
#         cleanup_old_files()
#         update_progress("starting", 5, "Saving files...")

#         # Get proper file extensions from uploaded files
#         ref_ext = get_file_extension(reference)
#         main_ext = get_file_extension(main)
        
#         # Save with proper extensions
#         ref_path = os.path.join(UPLOAD_DIR, f"ref_{session}{ref_ext}")
#         main_path = os.path.join(UPLOAD_DIR, f"main_{session}{main_ext}")
#         save_uploaded_file(reference, ref_path)
#         save_uploaded_file(main, main_path)

#         update_progress("reading", 15, "Reading data files...")
#         ref_df = read_saved_table(ref_path, sheet_name=ref_sheet)
#         main_df = read_saved_table(main_path, sheet_name=main_sheet)

#         def normalize(cols):
#             return (
#                 cols.astype(str)
#                 .str.strip()
#                 .str.replace("\ufeff", "", regex=False)
#                 .str.replace("\u200b", "", regex=False)
#                 .str.replace(r"\s+", " ", regex=True)
#                 .str.lower()
#             )

#         ref_df.columns = normalize(ref_df.columns)
#         main_df.columns = normalize(main_df.columns)

#         # Remove duplicate column headers
#         ref_df = ref_df.loc[:, ~ref_df.columns.duplicated()]
#         main_df = main_df.loc[:, ~main_df.columns.duplicated()]

#         ref_cols = list(ref_df.columns)
#         main_cols = list(main_df.columns)

#         # Column validation
#         missing_cols = [c for c in ref_cols if c not in main_cols]
#         extra_cols = [c for c in main_cols if c not in ref_cols]
#         if missing_cols or extra_cols:
#             update_progress("error", 100, "Column mismatch detected.")
#             return JSONResponse({
#                 "error": "Column mismatch",
#                 "missing_columns": missing_cols,
#                 "extra_columns": extra_cols
#             }, status_code=400)

#         shared_cols = [c for c in ref_cols if c in main_cols]
#         if not shared_cols:
#             return JSONResponse({"error": "No shared columns"}, status_code=400)

#         def norm_cell(x):
#             if pd.isna(x) or x is None:
#                 return ""
#             s = str(x).strip()
            
#             if not strict_mode:
#                 # Normalize whitespace
#                 s = " ".join(s.split())
                
#                 # Remove trailing .0 from numbers (e.g., "4219.0" -> "4219")
#                 if s.endswith('.0') and s[:-2].replace('.', '').replace('-', '').isdigit():
#                     s = s[:-2]
                
#                 # Try to normalize as number if possible
#                 try:
#                     # Check if it's a number
#                     float_val = float(s)
#                     # If it's a whole number, remove decimals
#                     if float_val == int(float_val):
#                         s = str(int(float_val))
#                     else:
#                         # Keep as float but normalize format
#                         s = str(float_val)
#                 except (ValueError, OverflowError):
#                     # Not a number, keep as string
#                     pass
            
#             return s

#         ref_df_proc = ref_df[shared_cols].fillna("").astype(str).applymap(norm_cell)
#         main_df_proc = main_df[shared_cols].fillna("").astype(str).applymap(norm_cell)

#         update_progress("comparing", 45, "Comparing rows...")
#         # Remove completely empty rows (all cells blank)
#         ref_df_proc = ref_df_proc.dropna(how="all")
#         main_df_proc = main_df_proc.dropna(how="all")

#         # Remove duplicates for clean comparison
#         ref_df_proc = ref_df_proc.drop_duplicates()
#         main_df_proc = main_df_proc.drop_duplicates()
        
#         # 🔄 SMART MATCHING: Instead of sorting, find best row matches
#         update_progress("matching", 50, "Finding matching rows...")
#         log(f"📊 Ref rows: {len(ref_df_proc)}, Main rows: {len(main_df_proc)}")
        
#         ref_tuples = [tuple(r) for r in ref_df_proc.values.tolist()]
#         main_tuples = [tuple(r) for r in main_df_proc.values.tolist()]
        
#         ref_counter = Counter(ref_tuples)
#         main_counter = Counter(main_tuples)
        
#         # Find exact matches first
#         exact_matches = {}
#         main_used = set()
        
#         for ref_idx, ref_tuple in enumerate(ref_tuples):
#             for main_idx, main_tuple in enumerate(main_tuples):
#                 if main_idx not in main_used and ref_tuple == main_tuple:
#                     exact_matches[ref_idx] = main_idx
#                     main_used.add(main_idx)
#                     break
        
#         # For remaining rows, find "close" matches (rows that differ in only a few cells)
#         ref_unmatched = [i for i in range(len(ref_tuples)) if i not in exact_matches]
#         main_unmatched = [i for i in range(len(main_tuples)) if i not in main_used]
        
#         close_matches = {}
#         similarity_threshold = len(shared_cols) * 0.7  # At least 70% columns must match
        
#         for ref_idx in ref_unmatched:
#             best_match = None
#             best_similarity = 0
            
#             for main_idx in main_unmatched:
#                 # Count matching cells
#                 matching_cells = sum(1 for k in range(len(shared_cols)) 
#                                    if ref_tuples[ref_idx][k] == main_tuples[main_idx][k])
                
#                 if matching_cells > best_similarity and matching_cells >= similarity_threshold:
#                     best_similarity = matching_cells
#                     best_match = main_idx
            
#             if best_match is not None:
#                 close_matches[ref_idx] = best_match
#                 main_used.add(best_match)
#                 main_unmatched.remove(best_match)
        
#         # Combine exact and close matches
#         all_matches = {**exact_matches, **close_matches}
#         matched_count = len(all_matches)
        
#         log(f"✅ Found {len(exact_matches)} exact matches, {len(close_matches)} close matches")

#         ref_tuples = [tuple(r) for r in ref_df_proc.values.tolist()]
#         main_tuples = [tuple(r) for r in main_df_proc.values.tolist()]

#         ref_counter = Counter(ref_tuples)
#         main_counter = Counter(main_tuples)

#         missing_rows, extra_rows = [], []

#         # Rows in reference but not matched
#         for i, ref_tuple in enumerate(ref_tuples):
#             if i not in all_matches:
#                 missing_rows.append(ref_tuple)
        
#         # Rows in main but not matched
#         for i, main_tuple in enumerate(main_tuples):
#             if i not in main_used:
#                 extra_rows.append(main_tuple)

#         missing_path = os.path.join(UPLOAD_DIR, f"missing_{session}.xlsx")
#         extra_path = os.path.join(UPLOAD_DIR, f"extra_{session}.xlsx")
#         mismatch_path = os.path.join(UPLOAD_DIR, f"mismatch_{session}.xlsx")

#         if missing_rows:
#             pd.DataFrame(missing_rows, columns=shared_cols).to_excel(missing_path, index=False)
#         if extra_rows:
#             pd.DataFrame(extra_rows, columns=shared_cols).to_excel(extra_path, index=False)

#         update_progress("analyzing", 70, "Analyzing cell-level mismatches...")
#         mismatches = []
#         mismatch_rows = []
        
#         # Check matched rows for cell-level differences
#         for ref_idx, main_idx in all_matches.items():
#             row_has_mismatch = False
#             for col_idx, col in enumerate(shared_cols):
#                 ref_val = ref_df_proc.iat[ref_idx, col_idx]
#                 main_val = main_df_proc.iat[main_idx, col_idx]
#                 if ref_val != main_val:
#                     row_has_mismatch = True
#                     mismatches.append({
#                         "ref_row": ref_idx,
#                         "main_row": main_idx,
#                         "column": col,
#                         "reference_value": ref_val,
#                         "main_value": main_val
#                     })
            
#             if row_has_mismatch:
#                 mismatch_rows.append((ref_idx, main_idx))
        
#         log(f"🔍 Found {len(mismatches)} cell mismatches in {len(mismatch_rows)} rows")

#         if mismatches:
#             if mismatch_rows:
#                 # Create complete row comparison with ALL columns
#                 combined_data = []
                
#                 for ref_idx, main_idx in mismatch_rows:
#                     # Add a row identifier
#                     ref_row = {"Source": "REFERENCE", "Ref Row": ref_idx + 2, "Main Row": main_idx + 2}
#                     main_row = {"Source": "MAIN", "Ref Row": ref_idx + 2, "Main Row": main_idx + 2}
                    
#                     # Add all columns from both reference and main
#                     for col in shared_cols:
#                         ref_val = ref_df_proc.iat[ref_idx, ref_df_proc.columns.get_loc(col)]
#                         main_val = main_df_proc.iat[main_idx, main_df_proc.columns.get_loc(col)]
#                         ref_row[col] = ref_val
#                         main_row[col] = main_val
                    
#                     combined_data.append(ref_row)
#                     combined_data.append(main_row)
                    
#                     # Add empty row for visual separation
#                     combined_data.append({col: "" for col in ["Source", "Ref Row", "Main Row"] + shared_cols})
                
#                 df_mismatch = pd.DataFrame(combined_data)
#                 df_mismatch.to_excel(mismatch_path, index=False)

#                 try:
#                     wb = load_workbook(mismatch_path)
#                     ws = wb.active
#                     red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
#                     blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
#                     green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

#                     # Build a map of which cells differ for each row pair
#                     mismatch_map = {}
#                     for m in mismatches:
#                         key = (m["ref_row"], m["main_row"])
#                         col = m["column"]
#                         if key not in mismatch_map:
#                             mismatch_map[key] = set()
#                         mismatch_map[key].add(col)

#                     # Color the cells: red for mismatches, green for matches
#                     row_num = 2  # Start after header
#                     for ref_idx, main_idx in mismatch_rows:
#                         mismatched_cols = mismatch_map.get((ref_idx, main_idx), set())
                        
#                         # Color REFERENCE row
#                         for col_idx, col_name in enumerate(["Source", "Ref Row", "Main Row"] + shared_cols, start=1):
#                             cell = ws.cell(row=row_num, column=col_idx)
#                             if col_name in mismatched_cols:
#                                 cell.fill = red
#                             elif col_name not in ["Source", "Ref Row", "Main Row"]:
#                                 cell.fill = green
#                             else:
#                                 cell.fill = blue
                        
#                         row_num += 1
                        
#                         # Color MAIN row
#                         for col_idx, col_name in enumerate(["Source", "Ref Row", "Main Row"] + shared_cols, start=1):
#                             cell = ws.cell(row=row_num, column=col_idx)
#                             if col_name in mismatched_cols:
#                                 cell.fill = red
#                             elif col_name not in ["Source", "Ref Row", "Main Row"]:
#                                 cell.fill = green
#                             else:
#                                 cell.fill = blue
                        
#                         row_num += 2  # Skip separator row

#                     wb.save(mismatch_path)
#                     wb.close()
#                 except Exception as e:
#                     log(f"⚠️ Failed to color mismatches: {e}")

#         try:
#             os.remove(ref_path)
#             os.remove(main_path)
#         except Exception:
#             pass

#         update_progress("done", 100, "Analysis complete ✅")

#         return {
#             "status": "success",
#             "session": session,
#             "rowsCompared": len(ref_df_proc),
#             "sharedColumns": len(shared_cols),
#             "matchedRows": matched_count,
#             "missingRows": len(missing_rows),
#             "extraRows": len(extra_rows),
#             "cellMismatches": len(mismatches),
#             "missingPath": os.path.basename(missing_path) if missing_rows else None,
#             "extraPath": os.path.basename(extra_path) if extra_rows else None,
#             "mismatchPath": os.path.basename(mismatch_path) if mismatches else None,
#         }

#     except Exception as e:
#         update_progress("error", 100, f"Error: {e}")
#         return JSONResponse({"error": str(e)}, status_code=500)


# # ----------------------------------------------------------
# # 🧩 Debug Samples
# # ----------------------------------------------------------
# @router.post("/debug/")
# async def debug_samples(reference: UploadFile = File(...), main: UploadFile = File(...), sample_size: int = 10):
#     """Return a small normalized sample for debugging."""
#     try:
#         # Get proper file extensions
#         ref_ext = get_file_extension(reference)
#         main_ext = get_file_extension(main)
        
#         ref_path = os.path.join(UPLOAD_DIR, f"debug_ref_{uuid.uuid4().hex[:6]}{ref_ext}")
#         main_path = os.path.join(UPLOAD_DIR, f"debug_main_{uuid.uuid4().hex[:6]}{main_ext}")
#         save_uploaded_file(reference, ref_path)
#         save_uploaded_file(main, main_path)

#         ref_df = read_saved_table(ref_path)
#         main_df = read_saved_table(main_path)

#         ref_df = ref_df.fillna("").astype(str)
#         main_df = main_df.fillna("").astype(str)
#         shared_cols = [c for c in ref_df.columns if c in main_df.columns]

#         return {
#             "reference_sample": ref_df[shared_cols].head(sample_size).values.tolist(),
#             "main_sample": main_df[shared_cols].head(sample_size).values.tolist(),
#             "shared_columns": shared_cols
#         }
#     except Exception as e:
#         return JSONResponse({"error": str(e)}, status_code=500)


# # ----------------------------------------------------------
# # 🟢 Progress Tracker
# # ----------------------------------------------------------
# @router.get("/progress")
# async def get_progress():
#     with progress_lock:
#         return ANALYSIS_PROGRESS.copy()


# # ----------------------------------------------------------
# # 📦 Download Endpoints
# # ----------------------------------------------------------
# def get_latest_file(prefix: str) -> Optional[str]:
#     files = [f for f in os.listdir(UPLOAD_DIR) if f.startswith(prefix)]
#     if not files:
#         return None
#     files.sort(key=lambda x: os.path.getmtime(os.path.join(UPLOAD_DIR, x)), reverse=True)
#     return os.path.join(UPLOAD_DIR, files[0])


# @router.get("/download-missing/")
# def download_missing(session: str = "latest"):
#     path = get_latest_file("missing_") if session == "latest" else os.path.join(UPLOAD_DIR, f"missing_{session}.xlsx")
#     if not path or not os.path.exists(path):
#         return JSONResponse({"error": "Missing rows file not found"}, status_code=404)
#     return FileResponse(path, filename="missing_rows.xlsx")


# @router.get("/download-extra/")
# def download_extra(session: str = "latest"):
#     path = get_latest_file("extra_") if session == "latest" else os.path.join(UPLOAD_DIR, f"extra_{session}.xlsx")
#     if not path or not os.path.exists(path):
#         return JSONResponse({"error": "Extra rows file not found"}, status_code=404)
#     return FileResponse(path, filename="extra_rows.xlsx")


# @router.get("/download-mismatched/")
# def download_mismatched(session: str = "latest"):
#     path = get_latest_file("mismatch_") if session == "latest" else os.path.join(UPLOAD_DIR, f"mismatch_{session}.xlsx")
#     if not path or not os.path.exists(path):
#         return JSONResponse({"error": "Mismatched rows file not found"}, status_code=404)
#     return FileResponse(path, filename="mismatched_rows.xlsx")

