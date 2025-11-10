from fastapi import APIRouter, UploadFile, File, HTTPException, Form
from fastapi.responses import JSONResponse, FileResponse
import pandas as pd
import os, shutil, json
from pathlib import Path
from typing import Optional
from charset_normalizer import from_bytes

router = APIRouter(tags=["Column Validator"])

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Persisted state within process (simple approach)
REFERENCE_PATH: Optional[str] = None
LAST_RESULT: Optional[dict] = None


# --------------------------------------------------------------------
# 🧩 Utility functions
# --------------------------------------------------------------------
def save_uploaded_file(upload: UploadFile, target_name: str) -> str:
    """Save UploadFile to uploads directory and return path."""
    path = os.path.join(UPLOAD_DIR, target_name)
    upload.file.seek(0)
    with open(path, "wb") as f:
        shutil.copyfileobj(upload.file, f)
    upload.file.seek(0)
    return path


def read_table(path: str, sheet_name: str = None) -> pd.DataFrame:
    """Read Excel or CSV into pandas DataFrame with auto header detection."""
    ext = Path(path).suffix.lower()
    try:
        if ext in (".xls", ".xlsx"):
            print(f"📖 Reading Excel file: {path}")
            if sheet_name:
                print(f"📄 Reading sheet: {sheet_name}")
            
            # Try openpyxl first with data_only to skip formulas/styles
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
                
                # Select the specified sheet or active sheet
                if sheet_name:
                    ws = wb[sheet_name]
                else:
                    ws = wb.active
                
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(list(row))
                wb.close()
                df = pd.DataFrame(data)
                print(f"✅ Successfully read with openpyxl (data_only mode)")
                print(f"📊 Raw data shape: {df.shape}")
                print(f"📊 First few rows: {df.head()}")
            except Exception as e1:
                print(f"⚠️ openpyxl data_only mode failed: {e1}")
                # Try standard pandas read
                try:
                    if sheet_name:
                        df = pd.read_excel(path, dtype=str, header=None, sheet_name=sheet_name)
                    else:
                        df = pd.read_excel(path, dtype=str, header=None)
                    print(f"✅ Successfully read with default engine")
                except Exception as e2:
                    print(f"❌ All engines failed")
                    raise Exception(f"Cannot read Excel file. The file may be corrupted. Please try: 1) Opening in Excel and saving as a new file, 2) Using 'Save As' and selecting .xlsx format. Error: {str(e1)}")
            
            print(f"✅ Excel file read successfully. Shape: {df.shape}")
            
            # find first non-empty header row
            header_found = False
            for i in range(min(5, len(df))):
                row = df.iloc[i]
                # Check if row has non-null, non-empty values
                if row.notna().any() and any(str(x).strip() for x in row if x is not None):
                    # Convert row to column names, handling None and converting to string
                    cols = []
                    for val in row:
                        if val is None or pd.isna(val):
                            cols.append("")
                        else:
                            cols.append(str(val).strip().replace("\ufeff", ""))
                    df.columns = cols
                    df = df.drop(index=i).reset_index(drop=True)
                    print(f"✅ Header found at row {i}: {df.columns.tolist()}")
                    header_found = True
                    break
            
            if not header_found:
                print(f"⚠️ No clear header found, using first row as header")
                if len(df) > 0:
                    df.columns = [str(x) if x is not None else "" for x in df.iloc[0]]
                    df = df.drop(index=0).reset_index(drop=True)
            
            return df
        else:
            # CSV
            print(f"📖 Reading CSV file: {path}")
            with open(path, "rb") as f:
                head = f.read(2048)
            try:
                result = from_bytes(head).best()
                enc = result.encoding if result else "utf-8"
            except Exception:
                enc = "utf-8"

            df = pd.read_csv(path, dtype=str, keep_default_na=False, encoding=enc, header=None)
            print(f"✅ CSV file read successfully. Shape: {df.shape}")
            
            # auto-detect header row
            for i in range(min(5, len(df))):
                if df.iloc[i].notna().any():
                    df.columns = df.iloc[i].astype(str).str.strip().str.replace("\ufeff", "")
                    df = df.drop(index=i).reset_index(drop=True)
                    print(f"✅ Header found at row {i}: {df.columns.tolist()}")
                    break
            return df
    except Exception as e:
        print(f"❌ Failed to read file: {path}")
        print(f"❌ Error: {str(e)}")
        raise ValueError(f"Failed to read table: {e}")


# --------------------------------------------------------------------
# 🟢 Upload Reference File
# --------------------------------------------------------------------
@router.post("/upload-reference/")
async def upload_reference(file: UploadFile = File(...)):
    """Upload and store a reference file. Returns columns and meta data."""
    global REFERENCE_PATH
    try:
        print(f"📥 Received file upload: {file.filename}")
        print(f"📥 Content type: {file.content_type}")
        
        filename = Path(file.filename).name
        print(f"📁 Saving as: reference_{filename}")
        
        saved_path = save_uploaded_file(file, f"reference_{filename}")
        print(f"✅ File saved to: {saved_path}")
        
        # Check if it's an Excel file and get sheet names
        ext = Path(saved_path).suffix.lower()
        sheets = []
        if ext in (".xls", ".xlsx"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(saved_path, data_only=True, read_only=True)
                sheets = wb.sheetnames
                wb.close()
                print(f"📄 Found {len(sheets)} sheet(s): {sheets}")
            except Exception as e:
                print(f"⚠️ Could not read sheet names: {e}")
        
        # Read the first sheet by default
        df = read_table(saved_path)
        
        if df is None or df.shape[1] == 0:
            raise ValueError("File has no columns")
        
        REFERENCE_PATH = saved_path
        columns = [str(c).strip() for c in df.columns.tolist()]
        
        print(f"✅ Reference file loaded: {filename}")
        print(f"✅ Columns detected: {columns}")
        
        return {
            "filename": filename,
            "columns": columns,
            "total_columns": len(columns),
            "uploadDate": pd.Timestamp.now().isoformat(),
            "sheets": sheets if sheets else None,
            "currentSheet": sheets[0] if sheets else None
        }
    except Exception as e:
        error_msg = str(e)
        print(f"❌ Upload failed: {error_msg}")
        raise HTTPException(status_code=400, detail=error_msg)


# --------------------------------------------------------------------
# 🔵 Change Sheet Selection
# --------------------------------------------------------------------
@router.post("/change-sheet/")
async def change_sheet(sheet_name: str):
    """Change the selected sheet for the reference file."""
    global REFERENCE_PATH
    try:
        if not REFERENCE_PATH:
            raise HTTPException(status_code=400, detail="No reference file uploaded")
        
        print(f"📄 Changing to sheet: {sheet_name}")
        
        # Read the specified sheet
        df = read_table(REFERENCE_PATH, sheet_name=sheet_name)
        
        if df is None or df.shape[1] == 0:
            raise ValueError(f"Sheet '{sheet_name}' has no columns")
        
        columns = [str(c).strip() for c in df.columns.tolist()]
        
        print(f"✅ Sheet changed to: {sheet_name}")
        print(f"✅ Columns detected: {columns}")
        
        return {
            "currentSheet": sheet_name,
            "columns": columns,
            "total_columns": len(columns)
        }
    except Exception as e:
        error_msg = str(e)
        print(f"❌ Failed to change sheet: {error_msg}")
        raise HTTPException(status_code=400, detail=error_msg)


# --------------------------------------------------------------------
# 🟡 Detect Sheets (for current file)
# --------------------------------------------------------------------
@router.post("/detect-sheets/")
async def detect_sheets(file: UploadFile = File(...)):
    """Detect available sheets in an Excel file without storing it."""
    try:
        filename = Path(file.filename).name
        ext = Path(filename).suffix.lower()
        
        if ext not in (".xls", ".xlsx"):
            return {"sheets": None, "currentSheet": None}
        
        # Save temporarily to read sheets
        temp_path = os.path.join(UPLOAD_DIR, f"temp_{filename}")
        temp_path = save_uploaded_file(file, f"temp_{filename}")
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(temp_path, data_only=True, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            print(f"📄 Detected {len(sheets)} sheet(s) in {filename}: {sheets}")
            
            # Clean up temp file
            os.remove(temp_path)
            
            return {
                "sheets": sheets,
                "currentSheet": sheets[0] if sheets else None
            }
        except Exception as e:
            print(f"⚠️ Could not read sheets: {e}")
            # Clean up temp file
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return {"sheets": None, "currentSheet": None}
            
    except Exception as e:
        error_msg = str(e)
        print(f"❌ Sheet detection failed: {error_msg}")
        raise HTTPException(status_code=400, detail=error_msg)


# --------------------------------------------------------------------
# 🟣 Compare Reference & Current Files
# --------------------------------------------------------------------
@router.post("/compare-reports/")
async def compare_reports(
    reference: UploadFile = File(...), 
    current: UploadFile = File(...),
    reference_sheet: Optional[str] = Form(None),
    current_sheet: Optional[str] = Form(None)
):
    """
    Compare uploaded reference & current files and return structured comparison.
    Expects multipart/form-data with keys: 'reference' and 'current'
    Optional: 'reference_sheet' and 'current_sheet' for Excel files
    """
    global LAST_RESULT, REFERENCE_PATH
    try:
        print(f"📊 Starting comparison...")
        print(f"📄 Reference: {reference.filename}")
        print(f"📄 Current: {current.filename}")
        print(f"📄 Reference sheet requested: {reference_sheet}")
        print(f"📄 Current sheet requested: {current_sheet}")
        
        # Save both files temporarily
        ref_name = Path(reference.filename).name
        cur_name = Path(current.filename).name
        ref_path = save_uploaded_file(reference, f"cmp_ref_{ref_name}")
        cur_path = save_uploaded_file(current, f"cmp_cur_{cur_name}")

        # Read dataframes with specified sheets if provided
        print(f"🔍 Reading reference file with sheet: {reference_sheet if reference_sheet else 'default (first sheet)'}")
        ref_df = read_table(ref_path, sheet_name=reference_sheet)
        
        print(f"🔍 Reading current file with sheet: {current_sheet if current_sheet else 'default (first sheet)'}")
        cur_df = read_table(cur_path, sheet_name=current_sheet)
        
        if reference_sheet:
            print(f"📄 Using reference sheet: {reference_sheet}")
        if current_sheet:
            print(f"📄 Using current sheet: {current_sheet}")

        # Normalize header strings
        ref_cols = [str(c).strip() for c in ref_df.columns.tolist()]
        cur_cols = [str(c).strip() for c in cur_df.columns.tolist()]

        # --- Missing / Extra ---
        missing = [c for c in ref_cols if c not in cur_cols]
        extra = [c for c in cur_cols if c not in ref_cols]

        # --- Case-only differences ---
        lower_ref_map = {c.lower(): c for c in ref_cols}
        case_diffs = []
        for c in cur_cols:
            cl = c.lower()
            if cl in lower_ref_map and lower_ref_map[cl] != c and c not in ref_cols:
                case_diffs.append({"expected": lower_ref_map[cl], "actual": c})

        # --- Sequence Errors ---
        sequence_errors = []
        for idx, c in enumerate(ref_cols):
            if c in cur_cols:
                exp_pos = idx + 1
                act_pos = cur_cols.index(c) + 1
                if exp_pos != act_pos:
                    sequence_errors.append({
                        "column": c,
                        "expectedPosition": exp_pos,
                        "actualPosition": act_pos
                    })

        # --- Build error list ---
        errors = []
        for c in missing:
            errors.append({"type": "missing", "message": f"Missing column: '{c}'", "expected": c})
        for c in extra:
            errors.append({"type": "extra", "message": f"Extra column: '{c}'", "actual": c})
        for cd in case_diffs:
            errors.append({
                "type": "case",
                "message": f"Case difference: expected '{cd['expected']}', actual '{cd['actual']}'",
                "expected": cd["expected"],
                "actual": cd["actual"]
            })
        for se in sequence_errors:
            errors.append({
                "type": "sequence",
                "message": f"Sequence mismatch for '{se['column']}': expected {se['expectedPosition']}, actual {se['actualPosition']}",
                "expected": se["column"],
                "actual": se["column"],
                "position": se["expectedPosition"],
                "actualPosition": se["actualPosition"]
            })

        summary = {
            "totalColumns": len(ref_cols),
            "matchingColumns": len(ref_cols) - len(missing),
            "missingColumns": len(missing),
            "extraColumns": len(extra),
            "sequenceErrors": len(sequence_errors)
        }

        result = {
            "isValid": len(errors) == 0,
            "summary": summary,
            "errors": errors,
            "referenceFile": Path(ref_path).name,
            "currentFile": Path(cur_path).name
        }

        LAST_RESULT = result

        with open(os.path.join(UPLOAD_DIR, "last_column_validation.json"), "w", encoding="utf-8") as f:
            json.dump(result, f, indent=2)

        print(f"✅ Comparison complete. Valid: {result['isValid']}")
        return result

    except Exception as e:
        error_msg = str(e)
        print(f"❌ Comparison failed: {error_msg}")
        return JSONResponse({"error": error_msg}, status_code=500)


# --------------------------------------------------------------------
# 🟠 Download Result JSON
# --------------------------------------------------------------------
@router.get("/download-result/")
def download_result():
    """Download the last comparison JSON result if it exists."""
    global LAST_RESULT
    try:
        if LAST_RESULT is None:
            return JSONResponse({"error": "No result available"}, status_code=404)
        path = os.path.join(UPLOAD_DIR, "last_column_validation.json")
        if not os.path.exists(path):
            with open(path, "w", encoding="utf-8") as f:
                json.dump(LAST_RESULT, f, indent=2)
        return FileResponse(
            path,
            filename="column_validation_result.json",
            media_type="application/json"
        )
    except Exception as e:
        return JSONResponse({"error": f"Failed to download result: {e}"}, status_code=500)
    
# from fastapi import APIRouter, UploadFile, File, HTTPException
# from fastapi.responses import JSONResponse, FileResponse
# import pandas as pd
# import os, shutil, json
# from pathlib import Path
# from typing import Optional
# from charset_normalizer import from_bytes  # ✅ Added missing import

# router = APIRouter(tags=["Column Validator"])

# UPLOAD_DIR = "uploads"
# os.makedirs(UPLOAD_DIR, exist_ok=True)

# # Persisted state within process (simple approach)
# REFERENCE_PATH: Optional[str] = None
# LAST_RESULT: Optional[dict] = None


# # --------------------------------------------------------------------
# # 🧩 Utility functions
# # --------------------------------------------------------------------
# def save_uploaded_file(upload: UploadFile, target_name: str) -> str:
#     """Save UploadFile to uploads directory and return path."""
#     path = os.path.join(UPLOAD_DIR, target_name)
#     upload.file.seek(0)
#     with open(path, "wb") as f:
#         shutil.copyfileobj(upload.file, f)
#     upload.file.seek(0)
#     return path


# def read_table(path: str) -> pd.DataFrame:
#     """Read Excel or CSV into pandas DataFrame with auto header detection."""
#     ext = Path(path).suffix.lower()
#     try:
#         if ext in (".xls", ".xlsx"):
#             df = pd.read_excel(path, dtype=str, engine="openpyxl", header=None)
#             # find first non-empty header row
#             for i in range(min(5, len(df))):
#                 if df.iloc[i].notna().any():
#                     df.columns = df.iloc[i].astype(str).str.strip().str.replace("\ufeff", "")
#                     df = df.drop(index=i).reset_index(drop=True)
#                     break
#             return df
#         else:
#             # CSV
#             with open(path, "rb") as f:
#                 head = f.read(2048)
#             try:
#                 result = from_bytes(head).best()
#                 enc = result.encoding if result else "utf-8"
#             except Exception:
#                 enc = "utf-8"

#             df = pd.read_csv(path, dtype=str, keep_default_na=False, encoding=enc, header=None)
#             # auto-detect header row
#             for i in range(min(5, len(df))):
#                 if df.iloc[i].notna().any():
#                     df.columns = df.iloc[i].astype(str).str.strip().str.replace("\ufeff", "")
#                     df = df.drop(index=i).reset_index(drop=True)
#                     break
#             return df
#     except Exception as e:
#         raise ValueError(f"Failed to read table: {e}")


# # --------------------------------------------------------------------
# # 🟢 Upload Reference File
# # --------------------------------------------------------------------
# @router.post("/upload-reference/")
# async def upload_reference(file: UploadFile = File(...)):
#     """Upload and store a reference file. Returns columns and meta data."""
#     global REFERENCE_PATH
#     try:
#         filename = Path(file.filename).name
#         saved_path = save_uploaded_file(file, f"reference_{filename}")
#         df = read_table(saved_path)
#         if df is None or df.shape[1] == 0:
#             raise ValueError("File has no columns")
#         REFERENCE_PATH = saved_path
#         columns = [str(c).strip() for c in df.columns.tolist()]
#         print(f"✅ Reference file loaded: {filename}")
#         print(f"✅ Columns detected: {columns}")
#         return {
#             "filename": filename,
#             "columns": columns,
#             "total_columns": len(columns),
#             "uploadDate": pd.Timestamp.now().isoformat()
#         }
#     except Exception as e:
#         raise HTTPException(status_code=400, detail=f"Failed to upload reference: {e}")


# # --------------------------------------------------------------------
# # 🟣 Compare Reference & Current Files
# # --------------------------------------------------------------------
# @router.post("/compare-reports/")
# async def compare_reports(reference: UploadFile = File(...), current: UploadFile = File(...)):
#     """
#     Compare uploaded reference & current files and return structured comparison.
#     Expects multipart/form-data with keys: 'reference' and 'current'
#     """
#     global LAST_RESULT, REFERENCE_PATH
#     try:
#         # Save both files temporarily
#         ref_name = Path(reference.filename).name
#         cur_name = Path(current.filename).name
#         ref_path = save_uploaded_file(reference, f"cmp_ref_{ref_name}")
#         cur_path = save_uploaded_file(current, f"cmp_cur_{cur_name}")

#         # Read dataframes
#         ref_df = read_table(ref_path)
#         cur_df = read_table(cur_path)

#         # Normalize header strings
#         ref_cols = [str(c).strip() for c in ref_df.columns.tolist()]
#         cur_cols = [str(c).strip() for c in cur_df.columns.tolist()]

#         # --- Missing / Extra ---
#         missing = [c for c in ref_cols if c not in cur_cols]
#         extra = [c for c in cur_cols if c not in ref_cols]

#         # --- Case-only differences ---
#         lower_ref_map = {c.lower(): c for c in ref_cols}
#         case_diffs = []
#         for c in cur_cols:
#             cl = c.lower()
#             if cl in lower_ref_map and lower_ref_map[cl] != c and c not in ref_cols:
#                 case_diffs.append({"expected": lower_ref_map[cl], "actual": c})

#         # --- Sequence Errors ---
#         sequence_errors = []
#         for idx, c in enumerate(ref_cols):
#             if c in cur_cols:
#                 exp_pos = idx + 1
#                 act_pos = cur_cols.index(c) + 1
#                 if exp_pos != act_pos:
#                     sequence_errors.append({
#                         "column": c,
#                         "expectedPosition": exp_pos,
#                         "actualPosition": act_pos
#                     })

#         # --- Build error list ---
#         errors = []
#         for c in missing:
#             errors.append({"type": "missing", "message": f"Missing column: '{c}'", "expected": c})
#         for c in extra:
#             errors.append({"type": "extra", "message": f"Extra column: '{c}'", "actual": c})
#         for cd in case_diffs:
#             errors.append({
#                 "type": "case",
#                 "message": f"Case difference: expected '{cd['expected']}', actual '{cd['actual']}'",
#                 "expected": cd["expected"],
#                 "actual": cd["actual"]
#             })
#         for se in sequence_errors:
#             errors.append({
#                 "type": "sequence",
#                 "message": f"Sequence mismatch for '{se['column']}': expected {se['expectedPosition']}, actual {se['actualPosition']}",
#                 "expected": se["column"],
#                 "actual": se["column"],
#                 "position": se["expectedPosition"],
#                 "actualPosition": se["actualPosition"]
#             })

#         summary = {
#             "totalColumns": len(ref_cols),
#             "matchingColumns": len(ref_cols) - len(missing),
#             "missingColumns": len(missing),
#             "extraColumns": len(extra),
#             "sequenceErrors": len(sequence_errors)
#         }

#         result = {
#             "isValid": len(errors) == 0,
#             "summary": summary,
#             "errors": errors,
#             "referenceFile": Path(ref_path).name,
#             "currentFile": Path(cur_path).name
#         }

#         LAST_RESULT = result

#         with open(os.path.join(UPLOAD_DIR, "last_column_validation.json"), "w", encoding="utf-8") as f:
#             json.dump(result, f, indent=2)

#         return result

#     except Exception as e:
#         print(f"❌ Comparison failed: {e}")
#         return JSONResponse({"error": str(e)}, status_code=500)


# # --------------------------------------------------------------------
# # 🟠 Download Result JSON
# # --------------------------------------------------------------------
# @router.get("/download-result/")
# def download_result():
#     """Download the last comparison JSON result if it exists."""
#     global LAST_RESULT
#     try:
#         if LAST_RESULT is None:
#             return JSONResponse({"error": "No result available"}, status_code=404)
#         path = os.path.join(UPLOAD_DIR, "last_column_validation.json")
#         if not os.path.exists(path):
#             with open(path, "w", encoding="utf-8") as f:
#                 json.dump(LAST_RESULT, f, indent=2)
#         return FileResponse(
#             path,
#             filename="column_validation_result.json",
#             media_type="application/json"
#         )
#     except Exception as e:
#         return JSONResponse({"error": f"Failed to download result: {e}"}, status_code=500)
